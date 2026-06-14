"""Leitura e parsing da planilha (read-only) + cálculo dos derivados.

Estratégia: ler apenas os *inputs crus* (reps, peso, veloc, PSE, tempo, bem-estar)
e calcular VTT e Carga U.A. em Python. Não confiamos no valor em cache das fórmulas
do Excel — em arquivo recém-editado ele vem ``None`` ou ``'#DIV/0!'``.
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import load_workbook

from . import cellmap as cm
from . import config
from .models import Exercise, Session, SetEntry, Week


# --------------------------------------------------------------------------- #
# helpers de limpeza de célula
# --------------------------------------------------------------------------- #
def _clean(value):
    """None / string vazia / erro de fórmula ('#DIV/0!') -> None."""
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        if not s or s.startswith("#"):
            return None
        return s
    return value


def _num(value) -> float | None:
    v = _clean(value)
    if isinstance(v, bool):  # bool é subclasse de int; não queremos
        return None
    if isinstance(v, (int, float)):
        return float(v)
    return None


def _date(value) -> str | None:
    v = _clean(value)
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    if isinstance(v, str):  # caso a DATA tenha sido digitada como texto
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
            try:
                return dt.datetime.strptime(v, fmt).date().isoformat()
            except ValueError:
                continue
    return None


# --------------------------------------------------------------------------- #
# parsing
# --------------------------------------------------------------------------- #
def _materialize(ws) -> dict:
    """Lê a aba (modo read-only) para um dict {(linha, coluna): valor}."""
    grid: dict[tuple[int, int], object] = {}
    for row in ws.iter_rows(
        min_row=1, max_row=cm.MAX_ROW, min_col=1, max_col=cm.MAX_COL
    ):
        for cell in row:
            if cell.value is not None:
                grid[(cell.row, cell.column)] = cell.value
    return grid


def parse_session(grid: dict, day_index: int) -> Session:
    base = cm.base_row(day_index)
    g = grid.get

    day = _clean(g((base + cm.OFF_DAY, cm.COL_DAYNAME))) or cm.DAY_NAMES[day_index]
    date = _date(g((base + cm.OFF_DAY, cm.COL_DATE)))

    wellness = {
        "sono": _num(g((base + cm.OFF_DAY, cm.COL_SONO))),
        "estresse": _num(g((base + cm.OFF_DAY, cm.COL_ESTRESSE))),
        "fadiga": _num(g((base + cm.OFF_DAY, cm.COL_FADIGA))),
        "dor": _num(g((base + cm.OFF_DAY, cm.COL_DOR))),
    }
    present = [v for v in wellness.values() if v is not None]
    readiness = round(sum(present) / len(present), 2) if present else None

    pse = _num(g((base + cm.OFF_PSE, cm.COL_PSE)))
    tempo = _num(g((base + cm.OFF_TEMPO, cm.COL_TEMPO)))
    carga_ua = round(pse * tempo, 1) if (pse is not None and tempo is not None) else None

    exercises: list[Exercise] = []
    for r in range(base + cm.OFF_EX_FIRST, base + cm.OFF_EX_LAST + 1):
        name = _clean(g((r, cm.COL_EX_NAME)))
        if not name:
            continue
        sets: list[SetEntry] = []
        ex_vtt = 0.0
        for i, (rc, pc, vc) in enumerate(cm.SERIES_COLS, start=1):
            reps = _num(g((r, rc)))
            peso = _num(g((r, pc)))
            veloc = _num(g((r, vc)))
            if reps is None and peso is None and veloc is None:
                continue
            vtt = (reps or 0.0) * (peso or 0.0)
            ex_vtt += vtt
            sets.append(SetEntry(i, reps, peso, veloc, round(vtt, 1)))
        if sets:
            exercises.append(Exercise(name, sets, round(ex_vtt, 1)))

    vtt_session = round(sum(e.vtt for e in exercises), 1)
    logged = any((s.peso or 0) > 0 for e in exercises for s in e.sets)
    planned = len(exercises) > 0

    return Session(
        day_index=day_index,
        day=day,
        date=date,
        exercises=exercises,
        pse=pse,
        tempo_min=tempo,
        carga_ua=carga_ua,
        vtt_session=vtt_session,
        wellness=wellness,
        readiness=readiness,
        logged=logged,
        planned=planned,
    )


def parse_week(grid: dict, week_number: int, sheet: str) -> Week:
    sessions = [parse_session(grid, d) for d in range(cm.N_DAYS)]
    return Week(week=week_number, sheet=sheet, sessions=sessions)


def load_weeks(path: str | Path | None = None) -> list[Week]:
    """Abre a planilha e devolve todas as semanas (abas ``SEM N``), ordenadas."""
    wb = load_workbook(path or config.xlsx_path(), data_only=True, read_only=True)
    try:
        weeks: list[Week] = []
        for title in wb.sheetnames:
            m = cm.WEEK_SHEET_RE.match(title)
            if not m:
                continue
            grid = _materialize(wb[title])
            weeks.append(parse_week(grid, int(m.group(1)), title))
        weeks.sort(key=lambda w: w.week)
        return weeks
    finally:
        wb.close()


def find_week(weeks: list[Week], week_number: int) -> Week:
    for w in weeks:
        if w.week == week_number:
            return w
    disponiveis = ", ".join(str(w.week) for w in weeks) or "(nenhuma)"
    raise ValueError(f"semana {week_number} não encontrada. Disponíveis: {disponiveis}")


# --------------------------------------------------------------------------- #
# resumos (funções puras sobre os dataclasses — testáveis sem MCP)
# --------------------------------------------------------------------------- #
def week_summary(w: Week) -> dict:
    days = []
    vtt_total = ua_total = 0.0
    logged = planned = 0
    readiness_vals = []
    dates = [s.date for s in w.sessions if s.date]
    for s in w.sessions:
        if s.planned:
            planned += 1
        if s.logged:
            logged += 1
            vtt_total += s.vtt_session
        if s.carga_ua is not None:
            ua_total += s.carga_ua
        if s.readiness is not None:
            readiness_vals.append(s.readiness)
        days.append(
            {
                "day": s.day,
                "day_index": s.day_index + 1,
                "date": s.date,
                "vtt_session": s.vtt_session,
                "carga_ua": s.carga_ua,
                "readiness": s.readiness,
                "logged": s.logged,
            }
        )
    return {
        "week": w.week,
        "sheet": w.sheet,
        "date_range": [min(dates), max(dates)] if dates else None,
        "sessions_logged": logged,
        "sessions_planned": planned,
        "vtt_total": round(vtt_total, 1),
        "carga_ua_total": round(ua_total, 1),
        "readiness_avg": round(sum(readiness_vals) / len(readiness_vals), 2)
        if readiness_vals
        else None,
        "days": days,
    }


def list_summary(weeks: list[Week]) -> list[dict]:
    out = []
    for w in weeks:
        s = week_summary(w)
        out.append(
            {
                "week": s["week"],
                "sheet": s["sheet"],
                "date_range": s["date_range"],
                "sessions_logged": s["sessions_logged"],
                "sessions_planned": s["sessions_planned"],
                "vtt_total": s["vtt_total"],
                "carga_ua_total": s["carga_ua_total"],
                "readiness_avg": s["readiness_avg"],
                "has_data": s["sessions_logged"] > 0,
            }
        )
    return out


def exercise_history(weeks: list[Week], query: str) -> dict:
    qn = cm.normalize(query)
    matched: set[str] = set()
    history = []
    for w in weeks:
        appearances = []
        max_peso = None
        total_vtt = total_reps = 0.0
        for s in w.sessions:
            for e in s.exercises:
                en = cm.normalize(e.name)
                if qn not in en and en not in qn:
                    continue
                matched.add(e.name)
                pesos = [st.peso for st in e.sets if st.peso]
                mp = max(pesos) if pesos else None
                if mp is not None:
                    max_peso = mp if max_peso is None else max(max_peso, mp)
                total_vtt += e.vtt
                total_reps += sum(st.reps for st in e.sets if st.reps)
                appearances.append(
                    {
                        "day": s.day,
                        "date": s.date,
                        "name": e.name,
                        "max_peso": mp,
                        "vtt": e.vtt,
                        "sets": [
                            {"serie": st.serie, "reps": st.reps, "peso": st.peso, "vtt": st.vtt}
                            for st in e.sets
                        ],
                    }
                )
        if appearances:
            history.append(
                {
                    "week": w.week,
                    "max_peso": max_peso,
                    "total_vtt": round(total_vtt, 1),
                    "total_reps": round(total_reps, 1),
                    "appearances": appearances,
                }
            )
    return {
        "query": query,
        "matched_names": sorted(matched),
        "weeks_found": len(history),
        "history": history,
    }
