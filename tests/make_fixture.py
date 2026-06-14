"""Gera a fixture sintética ``tests/fixtures/sample.xlsx`` a partir do template.

Usa o MESMO ``cellmap`` do leitor (para não divergir do mapa) — limpa os inputs de
todas as semanas e escreve algumas sessões com números conhecidos, usados nos asserts.

Uso:
    python tests/make_fixture.py [caminho_do_template.xlsx]

Sem argumento, usa ``config.xlsx_path()`` (env AMAZFITOPS_XLSX ou data/planilha.xlsx).
Como a fixture só precisa da ESTRUTURA do template (não de dados pessoais), o arquivo
gerado é seguro para versionar.
"""

from __future__ import annotations

import datetime as dt
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from amazfitops import cellmap as cm  # noqa: E402
from amazfitops import config  # noqa: E402

OUT = Path(__file__).resolve().parent / "fixtures" / "sample.xlsx"

# (semana, day_index) -> sessão.  Números escolhidos para asserts fáceis.
#   exercises: lista de (nome, [(reps, peso), ...])
FIXTURE = {
    (1, 0): {  # SEM 1, SEGUNDA -> VTT 4800, U.A. 7*50=350, prontidão 5.0
        "date": dt.datetime(2024, 11, 4),
        "wellness": (5, 5, 5, 5),
        "pse": 7, "tempo": 50,
        "exercises": [("Leg Press 45", [(12, 100), (12, 100), (12, 100), (12, 100)])],
    },
    (1, 5): {  # SEM 1, SÁBADO -> VTT 3200+1800=5000, U.A. 8*60=480, prontidão 4.0
        "date": dt.datetime(2024, 11, 9),
        "wellness": (4, 5, 4, 3),
        "pse": 8, "tempo": 60,
        "exercises": [
            ("Agachamento Livre", [(10, 80), (10, 80), (10, 80), (10, 80)]),
            ("Supino Reto", [(10, 60), (10, 60), (10, 60)]),
        ],
    },
    (2, 5): {  # SEM 2, SÁBADO -> agachamento progride p/ 85kg (VTT 3400), U.A. 8*65=520
        "date": dt.datetime(2024, 11, 16),
        "wellness": (5, 4, 4, 4),
        "pse": 8, "tempo": 65,
        "exercises": [("Agachamento Livre", [(10, 85), (10, 85), (10, 85), (10, 85)])],
    },
}


def _set(ws, row, col, value):
    ws[f"{get_column_letter(col)}{row}"] = value


def _clear_inputs(ws):
    for d in range(cm.N_DAYS):
        base = cm.base_row(d)
        for col in (cm.COL_DATE, cm.COL_SONO, cm.COL_ESTRESSE, cm.COL_FADIGA, cm.COL_DOR):
            _set(ws, base + cm.OFF_DAY, col, None)
        _set(ws, base + cm.OFF_PSE, cm.COL_PSE, None)
        _set(ws, base + cm.OFF_TEMPO, cm.COL_TEMPO, None)
        for r in range(base + cm.OFF_EX_FIRST, base + cm.OFF_EX_LAST + 1):
            for rc, pc, vc in cm.SERIES_COLS:
                _set(ws, r, rc, None)
                _set(ws, r, pc, None)
                _set(ws, r, vc, None)


def _write_session(ws, day_index, spec):
    base = cm.base_row(day_index)
    if spec.get("date"):
        _set(ws, base + cm.OFF_DAY, cm.COL_DATE, spec["date"])
    if spec.get("wellness"):
        sono, estresse, fadiga, dor = spec["wellness"]
        _set(ws, base + cm.OFF_DAY, cm.COL_SONO, sono)
        _set(ws, base + cm.OFF_DAY, cm.COL_ESTRESSE, estresse)
        _set(ws, base + cm.OFF_DAY, cm.COL_FADIGA, fadiga)
        _set(ws, base + cm.OFF_DAY, cm.COL_DOR, dor)
    if spec.get("pse") is not None:
        _set(ws, base + cm.OFF_PSE, cm.COL_PSE, spec["pse"])
    if spec.get("tempo") is not None:
        _set(ws, base + cm.OFF_TEMPO, cm.COL_TEMPO, spec["tempo"])
    for j, (name, sets) in enumerate(spec.get("exercises", [])):
        r = base + cm.OFF_EX_FIRST + j
        _set(ws, r, cm.COL_EX_NAME, name)
        for k, (reps, peso) in enumerate(sets):
            rc, pc, vc = cm.SERIES_COLS[k]
            _set(ws, r, rc, reps)
            _set(ws, r, pc, peso)
            _set(ws, r, vc, 2.0)


def build(template: str | Path | None = None) -> Path:
    src = Path(template) if template else config.xlsx_path()
    wb = load_workbook(src)  # leitura-escrita (não read-only)
    sheet_by_week = {}
    for title in wb.sheetnames:
        m = cm.WEEK_SHEET_RE.match(title)
        if m:
            sheet_by_week[int(m.group(1))] = title
    for title in sheet_by_week.values():
        _clear_inputs(wb[title])
    for (week, day_index), spec in FIXTURE.items():
        _write_session(wb[sheet_by_week[week]], day_index, spec)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    return OUT


if __name__ == "__main__":
    tpl = sys.argv[1] if len(sys.argv) > 1 else None
    print("fixture escrita em:", build(tpl))
